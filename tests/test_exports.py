from io import BytesIO

from django.test import TestCase
from django.contrib.auth import get_user_model

from lessons.models import LessonSession, UserSession, Classroom
from goals.models import Goal, KIInteraction, OverallGoal
from reflections.models import Reflection
import openpyxl


class ExportViewTests(TestCase):
    def setUp(self):
        User = get_user_model()
        self.client.force_login(User.objects.create_user(pseudonym="staff", password="pw", is_staff=True))
        self.class_a = Classroom.objects.create(name="10A")
        self.class_b = Classroom.objects.create(name="10B")
        self.lesson1 = LessonSession.objects.create(date="2024-01-01", classroom=self.class_a)
        self.lesson2 = LessonSession.objects.create(date="2024-05-01", classroom=self.class_a)
        self.lesson3 = LessonSession.objects.create(date="2024-01-01", classroom=self.class_b)
        self.user1 = User.objects.create_user(pseudonym="u1", gruppe=User.VG, classroom=self.class_a)
        self.user2 = User.objects.create_user(pseudonym="u2", gruppe=User.KG, classroom=self.class_b)
        self.session1 = UserSession.objects.create(user=self.user1, lesson_session=self.lesson1)
        self.session2 = UserSession.objects.create(user=self.user1, lesson_session=self.lesson2)
        self.session3 = UserSession.objects.create(user=self.user2, lesson_session=self.lesson3)
        self.goal1 = Goal.objects.create(user_session=self.session1, raw_text="g1")
        self.goal2 = Goal.objects.create(user_session=self.session2, raw_text="g2")
        self.goal3 = Goal.objects.create(user_session=self.session3, raw_text="g3")
        OverallGoal.objects.create(user=self.user1, text="old overall")
        OverallGoal.objects.create(user=self.user1, text="new overall")
        Reflection.objects.create(
            user_session=self.session2,
            goal=self.goal2,
            result="yes",
            obstacles="none",
            next_step="next",
            next_step_source="user",
        )
        KIInteraction.objects.create(goal=self.goal2, turn=1, role="user", content="hi")

    def test_csv_filters(self):
        resp = self.client.get("/api/export/csv/", {"from": "2024-02-01", "class": "10A", "group": "VG"})
        self.assertEqual(resp.status_code, 200)
        content = resp.content.decode()
        self.assertIn("g2", content)
        self.assertNotIn("g1", content)
        self.assertNotIn("g3", content)
        self.assertIn("new overall", content)

    def test_xlsx_contains_sheets_and_exported_at(self):
        resp = self.client.get("/api/export/xlsx/")
        self.assertEqual(resp.status_code, 200)
        wb = openpyxl.load_workbook(BytesIO(resp.content))
        self.assertEqual(
            set(wb.sheetnames),
            {"Users", "Goals", "Reflections", "KIInteractions", "OverallGoals", "flat_dataset"},
        )
        ws = wb["flat_dataset"]
        headers = [cell.value for cell in next(ws.iter_rows(max_row=1))]
        self.assertIn("exported_at", headers)
        self.assertIn("overall_goal", headers)
        data_row = [cell.value for cell in next(ws.iter_rows(min_row=2, max_row=2))]
        self.assertIn("new overall", data_row)
        ws_goals = wb["OverallGoals"]
        og_headers = [cell.value for cell in next(ws_goals.iter_rows(max_row=1))]
        self.assertIn("exported_at", og_headers)
